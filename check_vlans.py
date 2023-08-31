import paramiko
import os
import openpyxl
import time

host = "XXX.XXX.XXX.XXX"
username = "username"
password = "password"

client = paramiko.client.SSHClient()
client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
client.connect(host, username=username, password=password)

#opens the file that contains the list of vlans to be checked
dataframe = openpyxl.load_workbook("C:\\XXXXXX\\XXXXXX\\XXXXXX\\XXXXXX\\XXXXXX\\list_vlans.xlsx")

dataframe1 = dataframe.active

for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        vlan_to_check = col[row].value
        print(vlan_to_check)

        command = "show vlan id " + str(vlan_to_check)
        client.connect(host, username=username, password=password)
        _stdin, _stdout,_stderr = client.exec_command(command)
        output = _stdout.read()

        output2 = str(output)

        if output2.find("Po100") != -1:
            file1 = open("C:\\XXXXXX\\XXXXXX\\XXXXXX\\XXXXXX\\XXXXXX\\Vlans_resultsFE.txt", "a")
            file1.write(str(vlan_to_check) + " in CORE with PO100" + '\n')
            file1.close()
        elif output2.find("not found in current VLAN database") != -1:
            file1 = open("C:\\XXXXXX\\XXXXXX\\XXXXXX\\XXXXXX\\XXXXXX\\\\Vlans_resultsFE.txt", "a")
            file1.write(str(vlan_to_check) + " not in CORE" + '\n')
            file1.close()
        else:
            file1 = open("C:\\XXXXXX\\XXXXXX\\XXXXXX\\XXXXXX\\XXXXXX\\\\Vlans_resultsFE.txt", "a")
            file1.write(str(vlan_to_check) + " in CORE but no PO100" + '\n')
            file1.close()
    time.sleep(1)

        
client.close()
